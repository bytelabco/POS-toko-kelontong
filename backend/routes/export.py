from flask import Blueprint, make_response, request
from flask_jwt_extended import jwt_required
from models.models import Transaction, TransactionItem, Product
from routes.auth import owner_required
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import io

export_bp = Blueprint('export', __name__)

def get_filtered_transactions():
    date_from  = request.args.get('date_from')
    date_to    = request.args.get('date_to')
    status     = request.args.get('status')

    query = Transaction.query

    if date_from:
        query = query.filter(Transaction.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        dt_to = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Transaction.created_at < dt_to)
    if status and status != 'all':
        query = query.filter(Transaction.status == status)

    transactions = query.order_by(Transaction.created_at.desc()).all()

    result = []
    for t in transactions:
        items = TransactionItem.query.filter_by(transaction_id=t.id).all()
        item_names = ', '.join([
            f"{Product.query.get(i.product_id).name} x{i.quantity}"
            for i in items
        ])
        result.append({
            'id':             t.id,
            'created_at':     t.created_at.strftime('%d/%m/%Y %H:%M'),
            'items':          item_names,
            'total_price':    t.total_price,
            'payment_method': t.payment_method.upper() if t.payment_method else 'CASH',
            'status':         'Selesai' if t.status == 'completed' else 'Dibatalkan'
        })
    return result

@export_bp.route('/api/export/excel', methods=['GET'])
@owner_required
def export_excel():
    data = get_filtered_transactions()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Transaksi"

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers    = ['#', 'Tanggal', 'Item', 'Total', 'Metode Bayar', 'Status']
    col_widths = [5, 20, 40, 15, 15, 12]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30

    for row, t in enumerate(data, 2):
        values = [t['id'], t['created_at'], t['items'],
                  t['total_price'], t['payment_method'], t['status']]
        for col, value in enumerate(values, 1):
            cell           = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col == 4:
                cell.number_format = '#,##0'
            if t['status'] == 'Dibatalkan':
                cell.font = Font(color="EF4444")
        ws.row_dimensions[row].height = 20

    total_row   = len(data) + 2
    ws.cell(row=total_row, column=3, value="TOTAL PENDAPATAN").font = Font(bold=True)
    total_value = sum(t['total_price'] for t in data if t['status'] == 'Selesai')
    total_cell  = ws.cell(row=total_row, column=4, value=total_value)
    total_cell.font         = Font(bold=True)
    total_cell.number_format = '#,##0'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"laporan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = make_response(output.read())
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

@export_bp.route('/api/export/pdf', methods=['GET'])
@owner_required
def export_pdf():
    data   = get_filtered_transactions()
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               rightMargin=2*cm, leftMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)

    styles  = getSampleStyleSheet()
    content = []

    content.append(Paragraph("<b>LAPORAN TRANSAKSI</b>", styles['Title']))
    content.append(Paragraph(
        f"Digenerate: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    ))
    content.append(Spacer(1, 0.5*cm))

    total_pendapatan = sum(t['total_price'] for t in data if t['status'] == 'Selesai')
    total_void       = sum(1 for t in data if t['status'] == 'Dibatalkan')

    summary_data = [
        ['Total Transaksi',       str(len(data))],
        ['Total Pendapatan',      f"Rp {total_pendapatan:,.0f}"],
        ['Transaksi Dibatalkan',  str(total_void)],
    ]
    summary_table = Table(summary_data, colWidths=[5*cm, 5*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F1F5F9')),
        ('FONTNAME',   (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE',   (0,0), (-1,-1), 10),
        ('PADDING',    (0,0), (-1,-1), 6),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
    ]))
    content.append(summary_table)
    content.append(Spacer(1, 0.5*cm))

    table_data = [['#', 'Tanggal', 'Item', 'Total', 'Metode', 'Status']]
    for t in data:
        table_data.append([
            f"#{t['id']}",
            t['created_at'],
            t['items'][:40] + '...' if len(t['items']) > 40 else t['items'],
            f"Rp {t['total_price']:,.0f}",
            t['payment_method'],
            t['status']
        ])

    table = Table(table_data, colWidths=[1*cm, 3.5*cm, 5.5*cm, 3*cm, 2*cm, 2.5*cm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',     (0,0),  (-1,0),  colors.HexColor('#0F172A')),
        ('TEXTCOLOR',      (0,0),  (-1,0),  colors.white),
        ('FONTNAME',       (0,0),  (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0,0),  (-1,-1), 8),
        ('PADDING',        (0,0),  (-1,-1), 5),
        ('GRID',           (0,0),  (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0,1),  (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('ALIGN',          (3,0),  (3,-1),  'RIGHT'),
    ]))
    content.append(table)

    doc.build(content)
    buffer.seek(0)

    filename = f"laporan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response